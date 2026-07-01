from __future__ import annotations

import csv
import os
import re
import shutil
import stat
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from math import ceil
from pathlib import Path
from urllib.parse import unquote, urlparse

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import ImageFont


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"
SUPPORTED_SUFFIXES = {".xlsx", ".xls", ".xlsm", ".csv", ".tsv"}
HEADER_ALIASES = {
    "交易时间": "交易时间",
    "交易完成时间": "交易时间",
    "销方名称": "销售企业",
    "销售企业名称": "销售企业",
    "商户编号": "商户编号",
    "核销商编": "商户编号",
    "经销商编号": "商户编号",
    "应收销售金额": "销售金额",
    "销售金额": "销售金额",
    "其他支付": "其他支付",
    "参考号": "交易参考号",
    "订单号": "交易订单号",
    "补贴销售金额": "补贴金额",
    "品类": "编码品类",
    "商品明细": "商品名称",
    "发票金额": "发票金额",
    "发票号": "发票号",
    "发票号码": "发票号",
    "ID": "ID",
}
APPLIANCE_DETAIL_HEADERS = (
    "拨付批次",
    "交易时间",
    "交易参考号",
    "交易订单号",
    "销售企业",
    "商户编号",
    "销售金额",
    "其他支付",
    "实收销售金额",
    "补贴金额",
    "补贴比例",
    "SN码",
    "所在地区",
    "商品编码",
    "能耗等级",
    "编码品类",
    "商品名称",
    "发票金额",
    "发票号",
    "ID",
)
DIGITAL_DETAIL_HEADERS = (
    "拨付批次",
    "交易时间",
    "交易参考号",
    "交易订单号",
    "销售企业",
    "商户编号",
    "销售金额",
    "其他支付",
    "实收销售金额",
    "补贴金额",
    "编码品类",
    "商品名称",
    "SN码",
    "IMEI1码",
    "IMEI2码",
    "商品编码",
    "所在地区",
    "发票号",
    "备注",
)
DERIVED_HEADERS = ("财务大类", "品牌")
APPLIANCE_CATEGORY_MAP = {
    "A04-空调": "空调",
    "A05-电脑": "电脑",
    "A06-热水器": "厨卫",
    "A01-电视机": "电视",
    "A03-洗衣机": "洗衣机",
    "A02-电冰箱": "冰箱",
}
DIGITAL_CATEGORY_MAP = {
    "B01-手机": "手机",
    "B02-平板": "平板",
    "B03-智能手表手环": "智能穿戴",
}
APPLIANCE_BRAND_KEYWORDS = (
    ("华为", ("华为", "HUAWEI", "MATEBOOK")),
    ("小米", ("小米", "XIAOMI", "REDMI", "米家", "MIJIA")),
    ("A.O.史密斯", ("A.O.史密斯", "AO史密斯")),
    ("卡萨帝", ("卡萨帝", "CASARTE")),
    ("小天鹅", ("小天鹅",)),
    ("西门子", ("西门子",)),
    ("奥克斯", ("奥克斯",)),
    ("万家乐", ("万家乐",)),
    ("Leader", ("LEADER",)),
    ("COLMO", ("COLMO",)),
    ("TCL", ("TCL",)),
    ("小鸭", ("小鸭",)),
    ("美菱", ("美菱",)),
    ("晶弘", ("晶弘",)),
    ("华凌", ("华凌",)),
    ("格力", ("格力",)),
    ("海尔", ("海尔",)),
    ("美的", ("美的", "MIDEA")),
    ("海信", ("海信",)),
    ("康佳", ("康佳", "KONKA")),
    ("三星", ("三星", "SAMSUNG")),
    ("索尼", ("索尼", "SONY")),
    ("LG", ("LG ",)),
    ("东芝", ("东芝", "TOSHIBA")),
    ("松下", ("松下", "PANASONIC")),
    ("长虹", ("长虹", "CHANGHONG")),
    ("容声", ("容声", "RONSHEN")),
    ("夏普", ("夏普", "SHARP")),
    ("飞利浦", ("飞利浦", "PHILIPS")),
    ("林内", ("林内", "RINNAI")),
    ("澳柯玛", ("澳柯玛", "AUCMA")),
    ("日立", ("日立", "HITACHI")),
    ("荣耀", ("荣耀", "HONOR")),
    ("苹果", ("苹果", "APPLE", "MACBOOK", "MAC MINI")),
    ("惠普", ("惠普", "HP ", "OMEN ", "VICTUS ")),
    ("联想", ("联想", "LENOVO", "THINKBOOK", "小新")),
    ("华硕", ("华硕", "ASUS")),
    ("戴尔", ("戴尔", "DELL")),
    ("帅康", ("帅康", "SACON")),
    ("万和", ("万和", "VANWARD")),
    ("方太", ("方太", "FOTILE")),
    ("创维", ("创维",)),
    ("科龙", ("科龙",)),
    ("博世", ("博世",)),
    ("统帅", ("统帅",)),
)
DIGITAL_BRAND_KEYWORDS = (
    ("华为", ("华为", "HUAWEI", "PURA", "MATE", "NOVA")),
    ("小米", ("小米", "XIAOMI", "REDMI", "红米")),
    ("荣耀", ("荣耀", "HONOR")),
    ("OPPO", ("OPPO",)),
    (
        "vivo",
        (
            "VIVO",
            "IQOO",
            "X300",
            "X200",
            "Y600",
            "Y500",
            "Y300",
            "Y50 ",
            "S50",
            "S30",
        ),
    ),
    ("一加", ("一加", "ONEPLUS")),
    ("苹果", ("苹果", "IPHONE", "IPAD", "APPLE")),
    ("三星", ("三星", "SAMSUNG")),
    ("小天才", ("小天才",)),
    ("作业帮", ("作业帮",)),
    ("学而思", ("学而思",)),
    ("洪恩", ("洪恩",)),
    ("步步高", ("步步高",)),
)
APPLIANCE_BRAND_NORMALIZATION_MAP = {
    "COLMO": "美的",
    "卡萨帝": "海尔",
    "Leader": "海尔",
    "科龙": "海信",
    "统帅": "海尔",
    "华凌": "美的",
    "晶弘": "格力",
}
MIDEA_GROUP_CATEGORIES = {"洗衣机", "冰箱"}
MIDEA_GROUP_BRANDS = {"美的", "小天鹅", "东芝"}
APPLIANCE_BRAND_MODEL_ALIASES = {
    "TNDD20-08AIDE": "小天鹅",
    "65A6Q": "海信",
    "85A6Q": "海信",
    "KFR-50GW/CX2S": "美的",
    "NR-D521CG-W": "松下",
    "NR-D531AX-S": "松下",
    "NR-W592CG-S": "松下",
    "WV20G-H": "海信",
    "WV20W": "海信",
    "WH130U9Q-1": "海信",
    "WH130U9Q": "海信",
    "DSF-60D956": "帅康",
    "DSF-60J925": "帅康",
    "自然风 双出风": "小米",
    "自然风PRO 双出风": "小米",
    "柔风（1.5匹/变频/新一级能效）": "小米",
    "KFRD-72FW/C3AA1": "海信",
    "D60-H7C1": "万家乐",
    "D60-S3C2": "万家乐",
    "新风PRO 双出风": "小米",
    "超净洗 洗烘10KG银灰": "小米",
    "自然风（1.5匹/变频/一级能效）白色": "小米",
    "GMV舒韵家庭中央空调": "格力",
    "自然风（2匹//变频/新一级能效）": "小米",
    "DSF-60DY6(E)": "帅康",
    "D60-G5C1": "万家乐",
    "超净洗 滚筒10KG 银灰": "小米",
    "精护洗PRO 洗烘10KG": "小米",
}
MODEL_TOKEN_PATTERN = re.compile(
    r"(?=[A-Z0-9._/+\-]*\d)[A-Z0-9]+(?:[._/+\-][A-Z0-9]+)*",
    re.IGNORECASE,
)
MERGED_SHEET_NAME = "整合明细"
PREFERRED_FONT = "Maple Mono NF CN"
FALLBACK_FONT = "微软雅黑"
FONT_SIZE = 11
PIXEL_SAFETY_FACTOR = 1.25
SUMMARY_HEADERS = ["财务大类", "品牌", "补贴金额合计", "补贴金额计数"]


@dataclass(frozen=True)
class ProcessingProfile:
    name: str
    output_filename: str
    detail_headers: tuple[str, ...]
    optional_headers: frozenset[str]
    category_map: dict[str, str]
    brand_keywords: tuple[tuple[str, tuple[str, ...]], ...]
    brand_normalization_map: dict[str, str]
    brand_model_aliases: dict[str, str]


PROFILES = {
    "1": ProcessingProfile(
        name="家电",
        output_filename="家电回款明细.xlsx",
        detail_headers=APPLIANCE_DETAIL_HEADERS,
        optional_headers=frozenset({"发票金额"}),
        category_map=APPLIANCE_CATEGORY_MAP,
        brand_keywords=APPLIANCE_BRAND_KEYWORDS,
        brand_normalization_map=APPLIANCE_BRAND_NORMALIZATION_MAP,
        brand_model_aliases=APPLIANCE_BRAND_MODEL_ALIASES,
    ),
    "2": ProcessingProfile(
        name="数码",
        output_filename="数码回款明细.xlsx",
        detail_headers=DIGITAL_DETAIL_HEADERS,
        optional_headers=frozenset(),
        category_map=DIGITAL_CATEGORY_MAP,
        brand_keywords=DIGITAL_BRAND_KEYWORDS,
        brand_normalization_map={},
        brand_model_aliases={},
    ),
}


def _convert_delimited(source: Path, destination: Path) -> None:
    raw = source.read_bytes()
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError(f"无法识别文本编码：{source}")

    delimiter = "\t" if source.suffix.lower() == ".tsv" else ","
    if source.suffix.lower() == ".csv":
        try:
            delimiter = csv.Sniffer().sniff(text[:8192]).delimiter
        except csv.Error:
            pass

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Sheet1")
    for row in csv.reader(text.splitlines(), delimiter=delimiter):
        worksheet.append(row)
    workbook.save(destination)


def _convert_xls(source: Path, destination: Path) -> None:
    source_book = xlrd.open_workbook(source)
    target_book = Workbook()
    target_book.remove(target_book.active)

    for source_sheet in source_book.sheets():
        target_sheet = target_book.create_sheet(source_sheet.name[:31])
        for row_index in range(source_sheet.nrows):
            values = []
            for cell in source_sheet.row(row_index):
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(value, source_book.datemode)
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    value = xlrd.error_text_from_code.get(value, f"错误代码 {value}")
                values.append(value)
            target_sheet.append(values)
    target_book.save(destination)


def _convert_xlsm(source: Path, destination: Path) -> None:
    # 输出格式固定为 xlsx，因此不保留 VBA 宏。
    workbook = load_workbook(source, keep_vba=False)
    workbook.save(destination)


def convert_source_to_xlsx(source: Path) -> Path:
    """Create one xlsx working copy; all later processing uses this path."""
    if not source.is_file():
        raise FileNotFoundError(f"原始数据文件不存在：{source}")
    if source.name.startswith("~$"):
        raise ValueError("不能选择 Excel 临时文件")
    suffix = source.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError(f"不支持的数据文件格式：{source.suffix or '无扩展名'}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    destination = OUTPUT_DIR / f".{source.stem}.working.xlsx"
    _remove_working_copy(destination)
    if suffix == ".xlsx":
        shutil.copyfile(source, destination)
    elif suffix == ".xls":
        _convert_xls(source, destination)
    elif suffix == ".xlsm":
        _convert_xlsm(source, destination)
    else:
        _convert_delimited(source, destination)
    print(f"已生成工作副本：{destination.relative_to(BASE_DIR)}")
    return destination


def _remove_working_copy(path: Path) -> None:
    if not path.exists():
        return
    path.chmod(path.stat().st_mode | stat.S_IWRITE)
    path.unlink()


def _iter_actual_rows(worksheet, max_column: int):
    """Yield actual rows and discard the source's trailing maximum-range formatting."""
    if hasattr(worksheet, "reset_dimensions"):
        worksheet.reset_dimensions()
    pending_empty: list[tuple] = []
    for row in worksheet.iter_rows(
        min_col=1, max_col=max_column, values_only=True
    ):
        values = tuple(row[:max_column])
        if any(value not in (None, "") for value in values):
            yield from pending_empty
            pending_empty.clear()
            yield values
        else:
            pending_empty.append(values)
            if len(pending_empty) >= 100:
                break


def _get_source_positions(
    row: tuple, sheet_name: str, profile: ProcessingProfile
) -> dict[str, int]:
    """Normalize and validate one source header row."""
    headers = [
        HEADER_ALIASES.get(value, value) if value is not None else ""
        for value in row
    ]
    header_counts = Counter(name for name in headers if name)
    duplicates = sorted(name for name, count in header_counts.items() if count > 1)
    if duplicates:
        raise ValueError(f"工作表 {sheet_name!r} 存在重复表头：{duplicates}")

    positions = {name: index for index, name in enumerate(headers) if name}
    missing = [
        name
        for name in profile.detail_headers
        if name not in profile.optional_headers and name not in positions
    ]
    if missing:
        raise ValueError(f"工作表 {sheet_name!r} 缺少字段：{missing}")
    return positions


def _write_normalized_detail(
    source,
    target,
    *,
    write_header: bool,
    profile: ProcessingProfile,
    merchant_id: str,
) -> tuple[int, int]:
    source_positions: dict[str, int] = {}
    written_rows = 0
    unidentified_brands = 0
    merchant_index = profile.detail_headers.index("商户编号")
    category_index = profile.detail_headers.index("编码品类")
    product_index = profile.detail_headers.index("商品名称")

    for row in _iter_actual_rows(source, max(20, len(profile.detail_headers))):
        if not source_positions and "拨付批次" in row:
            source_positions = _get_source_positions(row, source.title, profile)
            if write_header:
                target.append(profile.detail_headers + DERIVED_HEADERS)
            continue

        if not source_positions:
            continue

        normalized = [
            row[source_positions[name]] if name in source_positions else None
            for name in profile.detail_headers
        ]
        row_merchant_id = normalized[merchant_index]
        if str(row_merchant_id).strip() == merchant_id:
            encoded_category = normalized[category_index]
            financial_category = profile.category_map.get(encoded_category)
            if financial_category is None:
                raise ValueError(
                    f"工作表 {source.title!r} 存在未配置的编码品类："
                    f"{encoded_category!r}"
                )
            product_name = normalized[product_index]
            brand = _extract_brand(product_name, profile)
            brand = _normalize_financial_brand(brand, financial_category)
            if brand is None:
                unidentified_brands += 1
            target.append(normalized + [financial_category, brand])
            written_rows += 1

    if not source_positions:
        raise ValueError(f"工作表 {source.title!r} 未找到明细表头")
    return written_rows, unidentified_brands


def _extract_brand(
    product_name, profile: ProcessingProfile
) -> str | None:
    if product_name in (None, ""):
        return None
    normalized_name = str(product_name).strip().casefold()
    for brand, keywords in profile.brand_keywords:
        if any(keyword.casefold() in normalized_name for keyword in keywords):
            return profile.brand_normalization_map.get(brand, brand)
    for model, brand in profile.brand_model_aliases.items():
        if model.casefold() in normalized_name:
            return profile.brand_normalization_map.get(brand, brand)
    return None


def _normalize_financial_brand(
    brand: str | None, financial_category: str
) -> str | None:
    if (
        financial_category in MIDEA_GROUP_CATEGORIES
        and brand in MIDEA_GROUP_BRANDS
    ):
        return "美的系"
    return brand


def _extract_model_tokens(product_name) -> set[str]:
    if product_name in (None, ""):
        return set()
    return {
        token.upper().strip("._/+-")
        for token in MODEL_TOKEN_PATTERN.findall(str(product_name).upper())
        if len(token.strip("._/+-")) >= 6
    }


def _infer_missing_brands_from_existing_rows(worksheet) -> int:
    """Infer only when an exact model token maps to one known brand in this sheet."""
    header_positions = {
        cell.value: cell.column for cell in worksheet[1] if cell.value not in (None, "")
    }
    product_column = header_positions["商品名称"]
    brand_column = header_positions["品牌"]
    token_brands: dict[str, Counter] = defaultdict(Counter)

    for row in range(2, worksheet.max_row + 1):
        brand = worksheet.cell(row, brand_column).value
        if brand in (None, ""):
            continue
        for token in _extract_model_tokens(worksheet.cell(row, product_column).value):
            token_brands[token][brand] += 1

    unique_token_brand = {
        token: next(iter(counts))
        for token, counts in token_brands.items()
        if len(counts) == 1
    }
    inferred = 0
    for row in range(2, worksheet.max_row + 1):
        brand_cell = worksheet.cell(row, brand_column)
        if brand_cell.value not in (None, ""):
            continue
        candidates = {
            unique_token_brand[token]
            for token in _extract_model_tokens(
                worksheet.cell(row, product_column).value
            )
            if token in unique_token_brand
        }
        if len(candidates) == 1:
            brand_cell.value = candidates.pop()
            inferred += 1
    return inferred


def _select_output_font() -> tuple[str, Path | None]:
    """Return the requested font family and its regular font file when available."""
    if os.name == "nt":
        try:
            import winreg

            registry_path = r"SOFTWARE\Microsoft\Windows NT\CurrentVersion\Fonts"
            registered_fonts: list[tuple[str, str]] = []
            for hive in (winreg.HKEY_LOCAL_MACHINE, winreg.HKEY_CURRENT_USER):
                try:
                    with winreg.OpenKey(hive, registry_path) as key:
                        for index in range(winreg.QueryInfoKey(key)[1]):
                            name, filename, _ = winreg.EnumValue(key, index)
                            if isinstance(filename, str):
                                registered_fonts.append((name, filename))
                except OSError:
                    continue

            windows_fonts = Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts"

            preferred = [
                item
                for item in registered_fonts
                if item[0].casefold().startswith(PREFERRED_FONT.casefold())
            ]
            if preferred:
                regular = next(
                    (item for item in preferred if " regular " in f" {item[0].casefold()} "),
                    preferred[0],
                )
                font_path = Path(regular[1])
                if not font_path.is_absolute():
                    font_path = windows_fonts / font_path
                return PREFERRED_FONT, font_path if font_path.is_file() else None

            fallback = [
                item
                for item in registered_fonts
                if item[0].casefold().startswith("microsoft yahei")
                and "bold" not in item[0].casefold()
                and "light" not in item[0].casefold()
            ]
            if fallback:
                font_path = Path(fallback[0][1])
                if not font_path.is_absolute():
                    font_path = windows_fonts / font_path
                return FALLBACK_FONT, font_path if font_path.is_file() else None
        except (ImportError, OSError):
            pass
    return FALLBACK_FONT, None


def _pixel_width(value, font) -> float:
    if value is None:
        return 0.0
    return max((font.getlength(line) for line in str(value).splitlines()), default=0.0)


def _format_data_sheet(
    worksheet, font_name: str, font_path: Path | None
) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="FF000000")
    data_font = Font(name=font_name, size=FONT_SIZE)
    header_font = Font(name=font_name, size=FONT_SIZE, color="FFFFFFFF")
    data_alignment = Alignment(vertical="center", wrap_text=False)
    header_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=False
    )
    maximum_pixels = [0.0] * worksheet.max_column

    if font_path is not None:
        measurement_font = ImageFont.truetype(
            str(font_path), size=round(FONT_SIZE * 96 / 72)
        )
    else:
        measurement_font = ImageFont.load_default(size=round(FONT_SIZE * 96 / 72))

    worksheet.freeze_panes = "A2"
    for row_index, row in enumerate(worksheet.iter_rows(), 1):
        worksheet.row_dimensions[row_index].height = 20
        for column_index, cell in enumerate(row, 1):
            cell.font = header_font if row_index == 1 else data_font
            cell.alignment = header_alignment if row_index == 1 else data_alignment
            if row_index == 1:
                cell.fill = header_fill
            maximum_pixels[column_index - 1] = max(
                maximum_pixels[column_index - 1],
                _pixel_width(cell.value, measurement_font),
            )

    for column_index, pixels in enumerate(maximum_pixels, 1):
        # Excel column-width units are approximately seven pixels at 100% zoom.
        # The factor absorbs rendering differences between Pillow and Excel.
        width = ceil((pixels * PIXEL_SAFETY_FACTOR + 16) / 7)
        worksheet.column_dimensions[get_column_letter(column_index)].width = max(
            8, min(width, 255)
        )


def _build_summary_sheet(
    summary_sheet, detail_sheet, category_map: dict[str, str]
) -> int:
    header_positions = {
        cell.value: cell.column for cell in detail_sheet[1] if cell.value not in (None, "")
    }
    category_column = header_positions["财务大类"]
    brand_column = header_positions["品牌"]
    subsidy_column = header_positions["补贴金额"]
    groups: dict[tuple[str, str], list] = {}

    for row in range(2, detail_sheet.max_row + 1):
        category = detail_sheet.cell(row, category_column).value
        brand = detail_sheet.cell(row, brand_column).value
        subsidy = detail_sheet.cell(row, subsidy_column).value
        if category in (None, ""):
            raise ValueError(f"整合明细第 {row} 行缺少财务大类")
        key = (str(category), "" if brand in (None, "") else str(brand))
        if key not in groups:
            groups[key] = [Decimal("0"), 0]
        if subsidy not in (None, ""):
            try:
                subsidy_amount = Decimal(str(subsidy))
            except (InvalidOperation, ValueError) as error:
                raise ValueError(f"整合明细第 {row} 行补贴金额不是有效数值") from error
            groups[key][0] += subsidy_amount
            groups[key][1] += -1 if subsidy_amount < 0 else 1

    category_order = {
        category: index
        for index, category in enumerate(dict.fromkeys(category_map.values()))
    }
    summary_sheet.append(SUMMARY_HEADERS)
    for (category, brand), (amount, count) in sorted(
        groups.items(),
        key=lambda item: (
            category_order.get(item[0][0], len(category_order)),
            item[0][1],
        ),
    ):
        rounded_amount = amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        summary_sheet.append([category, brand, float(rounded_amount), count])
        summary_sheet.cell(summary_sheet.max_row, 3).number_format = "#,##0.00"
        summary_sheet.cell(summary_sheet.max_row, 4).number_format = "0"
    total_amount = sum((values[0] for values in groups.values()), Decimal("0"))
    total_count = sum(values[1] for values in groups.values())
    rounded_total = total_amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    summary_sheet.append(["合计", None, float(rounded_total), total_count])
    summary_sheet.cell(summary_sheet.max_row, 3).number_format = "#,##0.00"
    summary_sheet.cell(summary_sheet.max_row, 4).number_format = "0"
    return len(groups)


def _merge_summary_categories(summary_sheet) -> int:
    category_column = 1
    first_data_row = 2
    last_data_row = summary_sheet.max_row - 1  # Exclude the bottom total row.
    if last_data_row < first_data_row:
        return 0

    merged_groups = 0
    group_start = first_data_row
    current_value = summary_sheet.cell(group_start, category_column).value
    for row in range(first_data_row + 1, last_data_row + 2):
        next_value = (
            summary_sheet.cell(row, category_column).value
            if row <= last_data_row
            else object()
        )
        if next_value == current_value:
            continue
        group_end = row - 1
        if group_end > group_start:
            summary_sheet.merge_cells(
                start_row=group_start,
                start_column=category_column,
                end_row=group_end,
                end_column=category_column,
            )
            merged_groups += 1
        summary_sheet.cell(group_start, category_column).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        group_start = row
        current_value = next_value
    return merged_groups


def _replace_output(temporary: Path, final_path: Path) -> None:
    for attempt in range(5):
        try:
            os.replace(temporary, final_path)
            return
        except PermissionError as error:
            if attempt == 4:
                raise PermissionError(
                    f"无法更新 {final_path}，请先关闭 Excel/WPS 中打开的该文件"
                ) from error
            time.sleep(0.5)


def process_workbook(
    path: Path, profile: ProcessingProfile, merchant_id: str
) -> Path:
    """Stream normalized data from one working copy into the final workbook."""
    source_book = load_workbook(path, read_only=True, data_only=False)
    target_book = Workbook()
    target_book.remove(target_book.active)
    final_path = OUTPUT_DIR / profile.output_filename
    temporary = OUTPUT_DIR / f".{Path(profile.output_filename).stem}.tmp.xlsx"
    try:
        merged_sheet = None
        merged_rows = 0
        merged_header_written = False
        unidentified_brands = 0
        for source_sheet in source_book.worksheets:
            if source_sheet.title == "汇总":
                target_book.create_sheet(source_sheet.title)
            else:
                if merged_sheet is None:
                    merged_sheet = target_book.create_sheet(MERGED_SHEET_NAME)
                written_rows, missing_brands = _write_normalized_detail(
                    source_sheet,
                    merged_sheet,
                    write_header=not merged_header_written,
                    profile=profile,
                    merchant_id=merchant_id,
                )
                merged_rows += written_rows
                unidentified_brands += missing_brands
                merged_header_written = True
        if merged_sheet is None:
            raise ValueError(f"工作簿 {path.name!r} 中没有可整合的明细 Sheet")
        inferred_brands = _infer_missing_brands_from_existing_rows(merged_sheet)
        unidentified_brands -= inferred_brands
        if "汇总" not in target_book.sheetnames:
            summary_sheet = target_book.create_sheet("汇总", 0)
        else:
            summary_sheet = target_book["汇总"]
        summary_groups = _build_summary_sheet(
            summary_sheet, merged_sheet, profile.category_map
        )
        output_font, output_font_path = _select_output_font()
        _format_data_sheet(merged_sheet, output_font, output_font_path)
        _format_data_sheet(summary_sheet, output_font, output_font_path)
        merged_categories = _merge_summary_categories(summary_sheet)
        for cell in summary_sheet[summary_sheet.max_row]:
            cell.font = Font(name=output_font, size=FONT_SIZE, bold=True)
        target_book.save(temporary)
        source_book.close()
        _replace_output(temporary, final_path)
    finally:
        source_book.close()
        temporary.unlink(missing_ok=True)
        _remove_working_copy(path)
    print(
        f"已处理{profile.name}明细：{final_path.relative_to(BASE_DIR)}，"
        f"商户 {merchant_id} 共 {merged_rows} 条，"
        f"品牌推断 {inferred_brands} 条、未识别 {unidentified_brands} 条，"
        f"汇总 {summary_groups} 组、合并财务大类 {merged_categories} 组，"
        f"字体 {output_font}"
    )
    return final_path


def _parse_input_path(raw_path: str) -> Path:
    value = raw_path.strip().strip('"').strip("'")
    if value.casefold().startswith("file://"):
        parsed = urlparse(value)
        value = unquote(parsed.path)
        if os.name == "nt" and re.match(r"^/[A-Za-z]:", value):
            value = value[1:]
    return Path(value).expanduser().resolve()


def _prompt_for_merchant_id() -> str:
    while True:
        merchant_id = input("请输入要筛选的商户编号：").strip()
        if merchant_id:
            return merchant_id
        print("商户编号不能为空，请重新输入。")


def _prompt_for_job() -> tuple[ProcessingProfile, str, Path]:
    print("请选择数据类型：")
    print("1. 家电")
    print("2. 数码")
    choice = input("请输入序号：").strip()
    if choice not in PROFILES:
        raise ValueError("数据类型无效，请输入 1 或 2")
    merchant_id = _prompt_for_merchant_id()
    source = _parse_input_path(input("请输入原始数据文件地址："))
    return PROFILES[choice], merchant_id, source


def main() -> None:
    profile, merchant_id, source = _prompt_for_job()
    working_copy = convert_source_to_xlsx(source)
    final_path = process_workbook(working_copy, profile, merchant_id)
    print(f"处理完成：{final_path}")


if __name__ == "__main__":
    main()
